import re
import sys
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


try:
	from openpyxl import Workbook, load_workbook
	from openpyxl.utils import get_column_letter
except ImportError:
	Workbook = None
	get_column_letter = None
	load_workbook = None


WORKBOOK_PATTERN = "DUMP_Parshing_ALL SITE JABO CCSI & WPCList.xlsx"
DEFAULT_SHEET_NAMES = ("RETSUBNIT", "Display RET Subunit Dynamic Inf")

def find_workbook():
	"""Find the dump workbook next to this script."""
	script_folder = Path(__file__).resolve().parent
	workbook = script_folder / WORKBOOK_PATTERN
	if workbook.exists():
		return workbook

	candidates = sorted(
		path for path in script_folder.glob("*.xlsx") if not path.name.startswith("~$")
	)
	return candidates[0] if candidates else None


def site_tokens(value):
	"""Return all values enclosed by hash characters in a cell."""
	return {token.strip().casefold() for token in re.findall(r"#([^#]+)#", str(value))}


class DumpViewer(tk.Tk):
	def __init__(self, workbook_path):
		super().__init__()
		self.title("Huawei Dump Viewer Import and Export XLSX")
		self.geometry("1200x650")
		self.minsize(800, 450)
		self.configure(bg="#f4f1e8")
		style = ttk.Style(self)
		try:
			style.theme_use("clam")
		except tk.TclError:
			pass
		style.configure("App.TFrame", background="#f4f1e8")
		style.configure("Panel.TFrame", background="#fffdf7")
		style.configure("Title.TLabel", background="#fffdf7", foreground="#17221d", font=("Georgia", 20))
		style.configure("Subtitle.TLabel", background="#fffdf7", foreground="#64726a", font=("Segoe UI", 9))
		style.configure("Section.TLabel", background="#fffdf7", foreground="#176b4d", font=("Segoe UI", 9, "bold"))
		style.configure("Status.TLabel", background="#f4f1e8", foreground="#64726a", font=("Segoe UI", 9))
		style.configure("Accent.TButton", background="#176b4d", foreground="#ffffff", padding=(12, 8), font=("Segoe UI", 9, "bold"))
		style.map("Accent.TButton", background=[("active", "#0e5039")])
		style.configure("Treeview", rowheight=27, background="#fffdf7", fieldbackground="#fffdf7", font=("Segoe UI", 9))
		style.configure("Treeview.Heading", background="#e8efdb", foreground="#176b4d", font=("Segoe UI", 9, "bold"))
		self.workbook_path = workbook_path
		self.workbook = None
		self.headers = []
		self.rows = []

		self._build_widgets()
		self._load_workbook()

	def _build_widgets(self):
		controls = ttk.Frame(self, padding=(18, 16, 18, 10), style="App.TFrame")
		controls.pack(fill="x")
		panel = ttk.Frame(controls, padding=(16, 14), style="Panel.TFrame")
		panel.pack(fill="x")
		ttk.Label(panel, text="Mulai pencarian", style="Title.TLabel").grid(row=0, column=0, columnspan=8, sticky="w")
		tk.Label(panel, text="Import workbook, pilih sheet, lalu temukan data berdasarkan Site ID.", style="Subtitle.TLabel").grid(row=1, column=0, columnspan=8, sticky="w", pady=(2, 14))
		tk.Button(panel, text="Import File", command=self.import_workbook, style="Accent.TButton").grid(row=2, column=0, padx=(0, 8), sticky="ew")
		ttk.Button(panel, text="Export XLSX", command=self.export_xlsx).grid(row=2, column=1, padx=(0, 16), sticky="ew")

		tk.Label(panel, text="SHEET", style="Section.TLabel").grid(row=2, column=2, padx=(0, 6), sticky="e")
		self.sheet_name = tk.StringVar()
		self.sheet_box = ttk.Combobox(
			panel, textvariable=self.sheet_name, state="readonly", width=32
		)
		self.sheet_box.grid(row=2, column=3, padx=(0, 16), sticky="ew")
		self.sheet_box.bind("<<ComboboxSelected>>", lambda _event: self.search())

		tk.Label(panel, text="SITE ID", style="Section.TLabel").grid(row=2, column=4, padx=(0, 6), sticky="e")
		self.site_id = tk.StringVar()
		self.search_box = ttk.Entry(panel, textvariable=self.site_id, width=24)
		self.search_box.grid(row=2, column=5, padx=(0, 6), sticky="ew")
		self.search_box.bind("<Return>", lambda _event: self.search())
		ttk.Button(panel, text="Cari data", command=self.search, style="Accent.TButton").grid(row=2, column=6, padx=(0, 6), sticky="ew")
		ttk.Button(panel, text="Bersihkan", command=self.clear_results).grid(row=2, column=7, sticky="ew")
		for column in (3, 5):
			panel.columnconfigure(column, weight=1)

		self.status = tk.StringVar(value="Memuat workbook...")
		ttk.Label(self, textvariable=self.status, padding=(18, 6, 18, 10), style="Status.TLabel").pack(
			anchor="w"
		)

		table_frame = ttk.Frame(self, padding=(18, 0, 18, 16), style="App.TFrame")
		table_frame.pack(fill="both", expand=True)
		self.table = ttk.Treeview(table_frame, show="headings")
		y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.table.yview)
		x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.table.xview)
		self.table.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
		self.table.grid(row=0, column=0, sticky="nsew")
		y_scroll.grid(row=0, column=1, sticky="ns")
		x_scroll.grid(row=1, column=0, sticky="ew")
		table_frame.rowconfigure(0, weight=1)
		table_frame.columnconfigure(0, weight=1)

	def _load_workbook(self):
		if load_workbook is None:
			messagebox.showerror(
				"Dependensi belum ada",
				"Install dependensi terlebih dahulu dengan:\npython -m pip install openpyxl",
			)
			self.status.set("openpyxl belum terpasang")
			return
		if self.workbook_path is None:
			messagebox.showerror("File tidak ditemukan", "Workbook dump tidak ditemukan.")
			self.status.set("Workbook tidak ditemukan")
			return

		try:
			if self.workbook is not None:
				self.workbook.close()
			self.workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
			sheets = self.workbook.sheetnames
			self.sheet_box["values"] = sheets
			preferred = next((name for name in DEFAULT_SHEET_NAMES if name in sheets), sheets[0])
			self.sheet_name.set(preferred)
			self.status.set(f"File: {self.workbook_path.name} | Pilih site ID lalu tekan Cari")
		except Exception as error:
			messagebox.showerror("Gagal membaca workbook", str(error))
			self.status.set("Gagal membaca workbook")

	def import_workbook(self):
		selected_file = filedialog.askopenfilename(
			title="Pilih file dump Huawei",
			filetypes=[
				("Excel workbook", "*.xlsx *.xlsm"),
				("Semua file", "*.*"),
			],
		)
		if not selected_file:
			return

		self.workbook_path = Path(selected_file)
		self.site_id.set("")
		self.rows = []
		self._load_workbook()
		self._show_results()

	def search(self):
		if self.workbook is None or not self.site_id.get().strip():
			return

		query = self.site_id.get().strip().strip("#").casefold()
		sheet = self.workbook[self.sheet_name.get()]
		rows = sheet.iter_rows(values_only=True)
		try:
			self.headers = [str(value or "") for value in next(rows)]
		except StopIteration:
			self.headers = []
		self.rows = [
			tuple("" if value is None else str(value) for value in row)
			for row in rows
			if any(query in site_tokens(value) for value in row)
		]
		self._show_results()
		self.status.set(f"{len(self.rows)} baris ditemukan di sheet {self.sheet_name.get()}")

	def _show_results(self):
		self.table.delete(*self.table.get_children())
		columns = [f"column_{index}" for index in range(len(self.headers))]
		self.table["columns"] = columns
		for column_index, (column, header) in enumerate(zip(columns, self.headers)):
			self.table.heading(column, text=header)
			content_lengths = [
				len(row[column_index])
				for row in self.rows
				if column_index < len(row)
			]
			longest_value = max([len(header), *content_lengths], default=len(header))
			self.table.column(column, width=max(100, min(420, longest_value * 8 + 24)))
		for row in self.rows:
			self.table.insert("", "end", values=row)

	def clear_results(self):
		self.site_id.set("")
		self.rows = []
		self._show_results()
		self.status.set(f"File: {self.workbook_path.name if self.workbook_path else '-'}")

	def export_xlsx(self):
		if Workbook is None:
			messagebox.showerror(
				"Dependensi belum ada",
				"Install dependensi terlebih dahulu dengan:\npython -m pip install openpyxl",
			)
			return
		if not self.rows:
			messagebox.showinfo("Tidak ada data", "Lakukan pencarian sebelum export.")
			return

		sheet_part = self.sheet_name.get().strip() or "hasil"
		site_part = self.site_id.get().strip().strip("#") or "tanpa_site_id"
		default_name = re.sub(
			r'[<>:"/\\|?*]', "_", f"{sheet_part}_{site_part}"
		).strip(" .")
		output_file = filedialog.asksaveasfilename(
			title="Simpan hasil pencarian",
			initialfile=f"{default_name}.xlsx",
			defaultextension=".xlsx",
			filetypes=[("Excel workbook", "*.xlsx")],
		)
		if not output_file:
			return

		try:
			workbook_class = Workbook
			if workbook_class is None or get_column_letter is None:
				raise RuntimeError("Library openpyxl tidak tersedia.")
			workbook = workbook_class()
			worksheet = workbook.active
			if worksheet is None:
				raise RuntimeError("Worksheet export tidak dapat dibuat.")
			worksheet.title = self.sheet_name.get()[:31] or "Hasil"
			worksheet.append(self.headers)
			for row in self.rows:
				worksheet.append(row)
			worksheet.freeze_panes = "A2"
			worksheet.auto_filter.ref = worksheet.dimensions
			for column_index in range(1, worksheet.max_column + 1):
				column_letter = get_column_letter(column_index)
				column_cells = worksheet[column_letter]
				longest_value = max(
					len(str(cell.value or "")) for cell in column_cells
				)
				worksheet.column_dimensions[column_letter].width = max(
					10, min(50, longest_value + 2)
				)
			workbook.save(output_file)
			self.status.set(f"Export berhasil: {Path(output_file).name}")
		except Exception as error:
			messagebox.showerror("Export gagal", str(error))


def main():
	if sys.platform == "win32":
		try:
			from ctypes import windll

			windll.shcore.SetProcessDpiAwareness(1)
		except (ImportError, AttributeError, OSError):
			pass
	DumpViewer(find_workbook()).mainloop()


if __name__ == "__main__":
	main()
